import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference 
import matplotlib.pyplot as plt
 
#Create a new workbook
wb = Workbook()
 
#Create the first sheet
sheet1=wb.active 
sheet1.title = "Student Data"
 
# Populate the first sheet with data sheet1.append(["Student Name", "Roll No.", "Subject", "Marks"])
 
data = [
("Akash", 101, "Math", 90),
("vishwa", 102, "Math", 85),
("Nagendra", 103, "Math", 88),
("Akash", 101, "Science", 75),
("vishwa", 102, "Science", 80),
("Nagendra", 103, "Science", 82),
("Akash", 101, "English", 78),
("vishwa", 102, "English", 85),
("Nagendra", 103, "English", 80),
]
 
for row in data: sheet1.append(row)
 
# Create the second sheet for the graph
sheet2=wb.create_sheet(title="Average Marks")
 
#Calculate average marks for each subject
subjects = {}
for row in sheet1.iter_rows(min_row=2, max_row=sheet1.max_row, min_col=3, max_col=4, values_only=True):
    subject, marks = row
    if subject not in subjects:
        subjects[subject] = [marks]
    else:
        subjects[subject].append(marks)
 
# Calculate average for each subject
average_marks = {}
for subject, marks_list in subjects.items():
    average_marks[subject] = sum(marks_list) / len(marks_list)
 
# Populate data for the second sheet
sheet2.append(["Subject", "Average Marks"])
for subject, avg_mark in average_marks.items():
    sheet2.append([subject, avg_mark])
 
#Create bar chart in the second sheet
chart = BarChart()
chart.title = "Average Marks by Subject"
chart.x_axis.title = "Subject"
chart.y_axis.title = "Average Marks"
 
data = Reference(sheet2, min_col=2, min_row=2,
max_row=sheet2.max_row, max_col=2)
categories = Reference(sheet2, min_col=1, min_row=2,
max_row=sheet2.max_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories (categories)
sheet2.add_chart(chart, "C3")
 
#Save the workbook
wb.save("student_data.xlsx")
print("Excel file created successfully!")